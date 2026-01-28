import openpyxl
from openpyxl.cell.cell import MergedCell
import datetime 

def preparar_planilha(caminho_entrada, qtd_geradoras, qtd_beneficiarias):
    wb = openpyxl.load_workbook(caminho_entrada)
    
    # Preparar Geradoras
    if "UC GERADORA" in wb.sheetnames:
        ws_modelo_ger = wb["UC GERADORA"]
        for i in range(qtd_geradoras):
            nome_aba = f"UC GERADORA {i+1}" if i > 0 else "UC GERADORA"
            if i == 0: ws_modelo_ger.title = nome_aba
            else: 
                nova = wb.copy_worksheet(ws_modelo_ger)
                nova.title = nome_aba

    # Preparar Beneficiárias
    nome_modelo_benef = next((s for s in wb.sheetnames if "UC BENEF" in s.upper()), None)
    if nome_modelo_benef and qtd_beneficiarias > 0:
        ws_modelo_ben = wb[nome_modelo_benef]
        for i in range(qtd_beneficiarias):
            nome_aba = f"UC BENEF. {i+1}"
            if i == 0: ws_modelo_ben.title = nome_aba
            else: 
                nova = wb.copy_worksheet(ws_modelo_ben)
                nova.title = nome_aba

    return wb

def safe_write(ws, col, row, value):
    coord = f"{col}{row}"
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws[rng.start_cell.coordinate].value = value
                return
    else:
        cell.value = value

def salvar_dados_multiplos(wb, dados_estruturados):
    mapa_meses = {
        "JAN": "Jan", "FEV": "Fev", "MAR": "Mar", "ABR": "Abr",
        "MAI": "Mai", "JUN": "Jun", "JUL": "Jul", "AGO": "Ago",
        "SET": "Set", "OUT": "Out", "NOV": "Nov", "DEZ": "Dez"
    }

    # Definição das Colunas BASE
    cols = {
        'leitura_ant': 'B', 'leitura_atual': 'C',
        'geracao': 'I', 'credito': 'J',
        'consumo': 'K', 'valor': 'N',
        'saldo': 'P', # Padrão para Geradora
        'medidor': 'R', 'leitura_med_ant': 'S', 'leitura_med_atual': 'T'
    }

    for item in dados_estruturados:
        tipo = item['tipo']
        indice = item['indice']
        faturas = item['dados']
        
        # Nome da aba
        if tipo == 'geradora':
            nome_aba = "UC GERADORA" if indice == 1 and "UC GERADORA" in wb.sheetnames else f"UC GERADORA {indice}"
            col_saldo_atual = 'P' # Geradora usa P
            cols_uso = cols 
        else:
            nome_aba = f"UC BENEF. {indice}"
            col_saldo_atual = 'Q' # Beneficiária usa Q (SOLICITADO)
            cols_uso = {
            **cols,
            'consumo': 'F',
            'credito': 'H',
            'valor': 'J',
            'medidor': 'S',
            'leitura_med_ant': 'T', 
            'leitura_med_atual': 'U'
            }           
        if nome_aba in wb.sheetnames:
            ws = wb[nome_aba]

            for dados in faturas:
                # --- 1. DADOS DO MÊS ATUAL (DA FATURA) ---
                mes_pdf = dados.get("mes", "")
                if mes_pdf and mes_pdf in mapa_meses:
                    mes_excel = mapa_meses[mes_pdf]

                    # Achar linha
                    linha_destino = None
                    for row in range(5, 40):
                        celula = ws[f"A{row}"].value
                        if celula and str(celula).strip() == mes_excel:
                            linha_destino = row
                            break
                        
                    if linha_destino:
                        # Preenche tudo
                        ws[f"{cols['leitura_ant']}{linha_destino}"] = dados["data_leitura_anterior"]
                        ws[f"{cols['leitura_atual']}{linha_destino}"] = dados["data_leitura_atual"]
                        ws[f"{cols['geracao']}{linha_destino}"] = dados["energia_gerada"]
                        ws[f"{cols_uso['credito']}{linha_destino}"] = dados["credito_recebido"]
                        ws[f"{cols_uso['consumo']}{linha_destino}"] = dados["energia_ativa"]
                        ws[f"{cols_uso['valor']}{linha_destino}"] = dados["valor_fatura"]
                        ws[f"{col_saldo_atual}{linha_destino}"] = dados["saldo"] # P ou Q
                        ws[f"{cols_uso['medidor']}{linha_destino}"] = dados["medidor"]
                        ws[f"{cols_uso['leitura_med_ant']}{linha_destino}"] = dados["leitura_anterior"]
                        ws[f"{cols_uso['leitura_med_atual']}{linha_destino}"] = dados["leitura_atual"]

                # --- 2. PREENCHIMENTO RETROATIVO (HISTÓRICO) ---
                # Útil se enviou apenas 1 fatura e quer preencher os consumos anteriores
                if "historico" in dados and dados["historico"]:
                    for hist in dados["historico"]:
                        mes_hist = hist['mes']
                        linha_hist = None
                        for row in range(5, 45):
                            cell_val = ws[f"A{row}"].value
                            if not cell_val: continue
                            mes_comparacao = ""
                            # Se a célula for Data (comum no Excel), extraímos o mês dela
                            if isinstance(cell_val, (datetime.datetime, datetime.date)):
                                # Mapeia o número do mês (1-12) para a sigla (JAN-DEZ)
                                siglas = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
                                mes_comparacao = siglas[cell_val.month - 1]
                            else:
                                # Se for texto ("Jan", "Janeiro"), normalizamos para comparar
                                texto_cell = str(cell_val).strip().upper()
                                for sigla, nome_mapa in mapa_meses.items():
                                    if sigla in texto_cell or nome_mapa.upper() in texto_cell:
                                        mes_comparacao = sigla
                                        break
                            if mes_comparacao == mes_hist:
                                linha_hist = row
                                break
                        # Preenche o consumo se achar a linha e não for o mês da fatura atual
                        if linha_hist and mes_hist != dados.get("mes"):
                            col_cons = cols_uso['consumo']
                            ws[f"{col_cons}{linha_hist}"] = hist['consumo']
                                                
    # --- 3. RESUMO (UC e Endereço) ---
    ws_resumo = None
    for sheet in wb.sheetnames:
        if "RESUMO" in sheet.upper():
            ws_resumo = wb[sheet]
            break
    
    if ws_resumo:
        linha_atual = 7
        # Geradoras
        for item in dados_estruturados:
            if item['tipo'] == 'geradora' and item['dados']:
                dados_ref = item['dados'][0]
                safe_write(ws_resumo, "F", linha_atual, dados_ref.get("uc", ""))
                safe_write(ws_resumo, "G", linha_atual, dados_ref.get("endereco", ""))
                linha_atual += 1
        
        # Beneficiárias
        for item in dados_estruturados:
            if item['tipo'] == 'beneficiaria' and item['dados']:
                dados_ref = item['dados'][0]
                safe_write(ws_resumo, "F", linha_atual, dados_ref.get("uc", ""))
                safe_write(ws_resumo, "G", linha_atual, dados_ref.get("endereco", ""))
                linha_atual += 1
                
    return wb