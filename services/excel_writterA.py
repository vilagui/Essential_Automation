import openpyxl
from openpyxl.cell.cell import MergedCell
import datetime

def preparar_planilha(caminho_entrada, qtd_geradoras, qtd_beneficiarias):
    """Prepara o workbook duplicando as abas de modelo."""
    wb = openpyxl.load_workbook(caminho_entrada)
    
    # Preparar Geradoras
    if "UC GERADORA" in wb.sheetnames:
        ws_modelo_ger = wb["UC GERADORA"]
        for i in range(qtd_geradoras):
            nome_aba = f"UC GERADORA {i+1}" if i > 0 else "UC GERADORA"
            if i == 0: ws_modelo_ger.title = nome_aba
            else:
                nova = wb.copy_worksheet(ws_modelo_ger)
                nova.title = nome_aba

    # Preparar Beneficiárias
    nome_modelo_benef = next((s for s in wb.sheetnames if "UC BENEF" in s.upper()), None)
    if nome_modelo_benef and qtd_beneficiarias > 0:
        ws_modelo_ben = wb[nome_modelo_benef]
        for i in range(qtd_beneficiarias):
            nome_aba = f"UC BENEF. {i+1}"
            if i == 0: ws_modelo_ben.title = nome_aba
            else:
                nova = wb.copy_worksheet(ws_modelo_ben)
                nova.title = nome_aba
    return wb

def safe_write(ws, col, row, value):
    """Escreve em células, tratando corretamente células mescladas."""
    coord = f"{col}{row}"
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws[rng.start_cell.coordinate].value = value
                return
    else:
        cell.value = value

def salvar_dados_A(wb, dados_estruturados):
    """Mapeia os dados para as abas individuais, dimensionamento e resumo."""
    
    meses_map = {"JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6, 
                 "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12}

    nome_aba_geral = next((s for s in wb.sheetnames if "GRUPO A" in s.upper()), "GRUPO A")
    ws_geral = wb[nome_aba_geral] if nome_aba_geral in wb.sheetnames else None

    for item in dados_estruturados:
        tipo, indice, faturas = item['tipo'], item['indice'], item['dados']
        nome_aba_uc = "UC GERADORA" if tipo == 'geradora' and indice == 1 else (f"UC GERADORA {indice}" if tipo == 'geradora' else f"UC BENEF. {indice}")
        ws_uc = wb[nome_aba_uc] if nome_aba_uc in wb.sheetnames else None

        for dados in faturas:
            mes_num = meses_map.get(dados.get("mes"))
            if not mes_num: continue

            # --- 1. ABA DIMENSIONAMENTO GERAL ---
            if ws_geral:
                for row in range(5, 25):
                    celula = ws_geral[f"A{row}"].value
                    if isinstance(celula, datetime.datetime) and celula.month == mes_num:
                        # Dados consumo 
                        ws_geral[f"B{row}"] = dados.get("c_p", 0.0)
                        ws_geral[f"C{row}"] = dados.get("c_fp", 0.0)
                        ws_geral[f"D{row}"] = dados.get("c_hr", 0.0)
                        # Dados demanda 
                        ws_geral[f"M{row}"] = dados.get("d_p", 0.0)
                        ws_geral[f"N{row}"] = dados.get("d_fp", 0.0)
                        ws_geral[f"O{row}"] = dados.get("d_hr", 0.0)
                        break

            # --- 2. ABAS INDIVIDUAIS (Parte Amarela) ---
            if ws_uc:
                for row in range(5, 45):
                    celula = ws_uc[f"A{row}"].value
                    if isinstance(celula, datetime.datetime) and celula.month == mes_num:
                        ws_uc[f"B{row}"] = dados.get("data_leitura_anterior")
                        ws_uc[f"C{row}"] = dados.get("data_leitura_atual")
                        c_total = dados.get("c_p", 0) + dados.get("c_fp", 0) + dados.get("c_hr", 0)

                        if tipo == 'geradora':
                            ws_uc[f"I{row}"] = dados.get("energia_gerada", 0.0)
                            ws_uc[f"J{row}"] = dados.get("credito_recebido", 0.0)
                            ws_uc[f"N{row}"] = dados.get("valor_fatura", 0.0)
                            ws_uc[f"P{row}"] = dados.get("saldo", 0.0)
                        else:
                            ws_uc[f"F{row}"] = c_total
                            ws_uc[f"H{row}"] = dados.get("credito_recebido", 0.0)
                            ws_uc[f"J{row}"] = dados.get("valor_fatura", 0.0)
                            ws_uc[f"Q{row}"] = dados.get("saldo", 0.0)
                        break

    # --- 3. RESUMO (UC e Endereço) ---
    ws_resumo = next((wb[s] for s in wb.sheetnames if "RESUMO" in s.upper()), None)
    if ws_resumo:
        linha_atual = 7
        # Geradoras
        for item in dados_estruturados:
            if item['tipo'] == 'geradora' and item['dados']:
                dados_ref = item['dados'][0]
                safe_write(ws_resumo, "F", linha_atual, dados_ref.get("uc", ""))
                safe_write(ws_resumo, "G", linha_atual, dados_ref.get("endereco", ""))
                linha_atual += 1
        
        # Beneficiárias
        for item in dados_estruturados:
            if item['tipo'] == 'beneficiaria' and item['dados']:
                dados_ref = item['dados'][0]
                safe_write(ws_resumo, "F", linha_atual, dados_ref.get("uc", ""))
                safe_write(ws_resumo, "G", linha_atual, dados_ref.get("endereco", ""))
                linha_atual += 1
                
    return wb